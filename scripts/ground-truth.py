#!/usr/bin/env python3
"""Independent ground-truth extraction from the SCF workbook (openpyxl, NOT SheetJS).

Mirrors the app's documented parsing conventions so any disagreement between the
two independent parsers is a data-integrity finding.

Usage: SCF_XLSX=/path/to/scf.xlsx python3 scripts/ground-truth.py > ground-truth.json
Requires: openpyxl
"""
import hashlib
import json
import os
import re
import sys

import openpyxl

SRC = os.environ.get(
    "SCF_XLSX",
    "/home/mark/scf-releases/2026.1.1/Secure.Controls.Framework.SCF.-.2026.1.1.xlsx",
)

SAMPLE_IDS = [
    "GOV-01", "AST-01", "IAC-01", "NET-01", "CRY-01",
    "DCH-01", "HRS-01", "IRO-01", "RSK-01", "TPM-01",
]
SAMPLE_FRAMEWORKS = ["nist-800-53-r5", "iso-27002-2022", "emea-eu-nis2"]


def norm(v):
    return re.sub(r"\s+", " ", str(v if v is not None else "")).strip()


def slugify(h):
    return re.sub(r"^-|-$", "", re.sub(r"[^a-z0-9]+", "-", norm(h).lower().replace("+", " plus")))


def split_multi(v):
    return [x.strip() for x in str(v if v is not None else "").split("\n") if x.strip()]


def md5(s):
    return hashlib.md5(s.encode()).hexdigest()


CORE = [
    (r"^scf domain$", "domain"),
    (r"^scf control$", "name"),
    (r"^scf #$", "id"),
    (r"^secure controls framework \(scf\) control description$", "description"),
    (r"^conformity validation cadence$", "cadence"),
    (r"^evidence request list \(erl\) #$", "erl"),
    (r"^scf control question$", "question"),
    (r"^relative control weighting$", "weighting"),
    (r"^pptdf applicability$", "pptdf"),
    (r"^nist csf function grouping$", "csf"),
]
SKIPPED = r"^(minimum security requirements|identify (minimum compliance|discretionary security)|risk threat summary|control threat summary)"


def classify(header):
    h = norm(header)
    for pat, field in CORE:
        if re.match(pat, h, re.I):
            return ("core", field)
    m = re.match(r"^possible solutions & considerations (.+?) bls firm size", h, re.I)
    if m:
        return ("solution", m.group(1).strip())
    m = re.match(r"^scrm focus tier (\d)", h, re.I)
    if m:
        return ("tier", int(m.group(1)))
    m = re.match(r"^scr-cmm level (\d) (.+)$", h, re.I)
    if m:
        return ("maturity", int(m.group(1)))
    m = re.match(r"^scf (community derived|scrms|core .+)$", h, re.I)
    if m:
        return ("baseline", slugify(m.group(1)))
    m = re.match(r"^risk (r-[a-z]{2}-\d+)$", h, re.I)
    if m:
        return ("risk", m.group(1).upper())
    m = re.match(r"^threat ((?:nt|mt)-\d+)$", h, re.I)
    if m:
        return ("threat", m.group(1).upper())
    if re.match(r"^errata", h, re.I):
        return ("errata", None)
    if re.match(SKIPPED, h, re.I):
        return ("skipped", None)
    return ("framework", slugify(h))


wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)


def sheet_by(pattern):
    for name in wb.sheetnames:
        if re.search(pattern, name.strip(), re.I):
            return wb[name]
    raise KeyError(pattern)


out = {}

# ---- Main sheet -------------------------------------------------------------
ws = sheet_by(r"^scf 20")
rows = ws.iter_rows(values_only=True)
headers = next(rows)
kinds = [classify(h) if h is not None else ("skipped", None) for h in headers]

fw_cols = {}  # slug -> [col indexes]
for i, (kind, key) in enumerate(kinds):
    if kind == "framework":
        fw_cols.setdefault(key, []).append(i)
collisions = {k: v for k, v in fw_cols.items() if len(v) > 1}

core_idx = {key: i for i, (kind, key) in enumerate(kinds) if kind == "core"}

controls = 0
ids = []
weighting_sum = 0
weighting_null = 0
pptdf_total = 0
erl_refs_total = 0
risk_links = 0
threat_links = 0
maturity_nonempty = 0
solutions_nonempty = 0
baseline_counts = {}
tier_counts = {1: 0, 2: 0, 3: 0}
mapping_cells = 0
mapping_refs = 0
fwd_erl = {}
control_hashes = {}
per_fw_controls = {k: 0 for k in fw_cols}
samples = {}

for row in rows:
    cid = norm(row[core_idx["id"]])
    if not re.match(r"^[A-Z]{2,4}-\d", cid):
        continue
    controls += 1
    ids.append(cid)
    w = row[core_idx["weighting"]]
    try:
        wv = int(float(w))
        weighting_sum += wv
    except (TypeError, ValueError):
        weighting_null += 1
        wv = None
    pptdf = split_multi(row[core_idx["pptdf"]])
    pptdf_total += len(pptdf)
    erl_ids = split_multi(row[core_idx["erl"]])
    erl_refs_total += len(erl_ids)
    fwd_erl[cid] = set(erl_ids)
    row_fw = {}
    risk_ids, threat_ids = [], []
    maturity_texts = []
    solution_texts = []
    for i, (kind, key) in enumerate(kinds):
        v = row[i] if i < len(row) else None
        filled = v is not None and str(v).strip() != ""
        if kind == "framework" and filled:
            refs = split_multi(v)
            if refs:
                mapping_cells += 1
                mapping_refs += len(refs)
                row_fw[key] = refs
        elif kind == "risk" and filled:
            risk_links += 1
            risk_ids.append(key)
        elif kind == "threat" and filled:
            threat_links += 1
            threat_ids.append(key)
        elif kind == "maturity":
            maturity_texts.append((key, norm(v)))
            if filled:
                maturity_nonempty += 1
        elif kind == "solution" and filled:
            solutions_nonempty += 1
            solution_texts.append(norm(v))
        elif kind == "baseline" and filled:
            baseline_counts[key] = baseline_counts.get(key, 0) + 1
        elif kind == "tier" and filled:
            tier_counts[key] += 1
    for k in row_fw:
        per_fw_controls[k] += 1
    maturity_sorted = sorted(maturity_texts)
    control_hashes[cid] = md5(
        "\x1f".join(
            [
                cid,
                norm(row[core_idx["name"]]),
                norm(row[core_idx["description"]]),
                norm(row[core_idx["question"]]),
                norm(row[core_idx["cadence"]]),
                norm(row[core_idx["csf"]]),
                "" if wv is None else str(wv),
                ";".join(pptdf),
                ";".join(erl_ids),
                ";".join(sorted(risk_ids)),
                ";".join(sorted(threat_ids)),
                "|".join(t for _, t in maturity_sorted),
                "|".join(solution_texts),
                "|".join(
                    f"{fw}={';'.join(row_fw[fw])}" for fw in sorted(row_fw)
                ),
            ]
        )
    )
    if cid in SAMPLE_IDS:
        maturity_texts.sort()
        samples[cid] = {
            "description": str(row[core_idx["description"]] or "").strip(),
            "question": str(row[core_idx["question"]] or "").strip(),
            "cadence": norm(row[core_idx["cadence"]]),
            "csf": norm(row[core_idx["csf"]]),
            "weighting": wv,
            "pptdf": pptdf,
            "erlIds": erl_ids,
            "riskIds": sorted(risk_ids),
            "threatIds": sorted(threat_ids),
            "maturityHash": md5("|".join(t for _, t in maturity_texts)),
            "mappingFrameworkCount": len(row_fw),
            "mappings": {fw: row_fw.get(fw, []) for fw in SAMPLE_FRAMEWORKS},
        }

out["mainSheet"] = {
    "controls": controls,
    "idsHash": md5("|".join(sorted(ids))),
    "weightingSum": weighting_sum,
    "weightingNull": weighting_null,
    "pptdfTotal": pptdf_total,
    "erlRefsTotal": erl_refs_total,
    "riskLinks": risk_links,
    "threatLinks": threat_links,
    "maturityNonEmpty": maturity_nonempty,
    "solutionsNonEmpty": solutions_nonempty,
    "baselineCounts": baseline_counts,
    "tierCounts": {str(k): v for k, v in tier_counts.items()},
    "frameworkColumns": len(fw_cols),
    "slugCollisions": collisions,
    "mappingCells": mapping_cells,
    "mappingRefs": mapping_refs,
    "mappedFrameworks": sum(1 for v in per_fw_controls.values() if v > 0),
    "perFrameworkControls": per_fw_controls,
}
out["mainSheet"]["controlHashes"] = control_hashes
out["samples"] = samples

# ---- Domains ----------------------------------------------------------------
ws = sheet_by(r"domains & principles")
rows = list(ws.iter_rows(values_only=True))
hdr = [norm(h) for h in rows[0]]
cid = hdr.index("SCF Identifier")
dom_ids = [norm(r[cid]) for r in rows[1:] if norm(r[cid])]
out["domains"] = {"count": len(dom_ids), "ids": sorted(dom_ids)}

# ---- Sources ----------------------------------------------------------------
ws = sheet_by(r"authoritative sources")
rows = list(ws.iter_rows(values_only=True))
hdr = [norm(h) for h in rows[0]]
ch = hdr.index("SCF Column Header")
slugs = {slugify(r[ch]) for r in rows[1:] if norm(r[ch])}
out["sources"] = {"uniqueCount": len(slugs)}

# ---- Assessment objectives ---------------------------------------------------
ws = sheet_by(r"^assessment objectives")
rows = ws.iter_rows(values_only=True)
hdr = [norm(h) for h in next(rows)]
cao = next(i for i, h in enumerate(hdr) if re.match(r"^scf ao #$", h, re.I))
crig = next(i for i, h in enumerate(hdr) if re.match(r"^assessment rigor", h, re.I))
aos, rigor_nonnull = 0, 0
for r in rows:
    if norm(r[cao]):
        aos += 1
        try:
            if float(r[crig]):
                rigor_nonnull += 1
        except (TypeError, ValueError):
            pass
out["aos"] = {"count": aos, "rigorNonNull": rigor_nonnull}

# ---- ERL ---------------------------------------------------------------------
ws = sheet_by(r"^evidence request list")
rows = ws.iter_rows(values_only=True)
hdr = [norm(h) for h in next(rows)]
cid_ = hdr.index("ERL #")
cmap = hdr.index("SCF Control Mappings")
erl_count, erl_links = 0, 0
for r in rows:
    if norm(r[cid_]):
        erl_count += 1
        erl_links += len(split_multi(r[cmap]))
out["erl"] = {"count": erl_count, "controlLinks": erl_links}

# Evidence union (forward control->ERL refs + reverse ERL->control mappings)
ws = sheet_by(r"^evidence request list")
rows2 = ws.iter_rows(values_only=True)
hdr2 = [norm(h) for h in next(rows2)]
cid2, cm2 = hdr2.index("ERL #"), hdr2.index("SCF Control Mappings")
valid_erl, rev = set(), {}
for r in rows2:
    eid = norm(r[cid2])
    if not eid:
        continue
    valid_erl.add(eid)
    for c in split_multi(r[cm2]):
        rev.setdefault(c, set()).add(eid)
union_links = 0
for cid_x, f in fwd_erl.items():
    union_links += len((f & valid_erl) | rev.get(cid_x, set()))
for cid_x, rset in rev.items():
    if cid_x not in fwd_erl:
        union_links += len(rset)
out["erl"]["unionLinks"] = union_links
for sid, sample in out["samples"].items():
    sample["erlLinked"] = sorted(
        (set(sample["erlIds"]) & valid_erl) | rev.get(sid, set())
    )

# ---- Compensating ------------------------------------------------------------
ws = sheet_by(r"^compensating controls")
rows = ws.iter_rows(values_only=True)
hdr = [norm(h) for h in next(rows)]
cctl = hdr.index("SCF Control #")
groups = [i for i, h in enumerate(hdr) if re.match(r"^possible compensating control #\d", h, re.I)]
comp_count, comp_options = 0, 0
for r in rows:
    if not re.match(r"^[A-Z]{2,4}-\d", norm(r[cctl])):
        continue
    comp_count += 1
    for g in groups:
        oid = norm(r[g + 2]) if g + 2 < len(r) else ""
        if oid and not re.match(r"^n/?a$", oid, re.I):
            comp_options += 1
out["compensating"] = {"count": comp_count, "options": comp_options}

# ---- Privacy principles --------------------------------------------------------
ws = sheet_by(r"data privacy mgmt principles")
rows = ws.iter_rows(values_only=True)
hdr = [norm(h) for h in next(rows)]
cnum = hdr.index("#")
cctl = next(i for i, h in enumerate(hdr) if re.match(r"^scf #$", h, re.I))
principles = {}
last = ""
for r in rows:
    raw = r[cnum]
    if raw is not None and str(raw).strip():
        if isinstance(raw, float) and raw.is_integer():
            last = str(int(raw))
        else:
            last = str(raw).strip()
    n = last
    if n == "":
        continue
    c = norm(r[cctl])
    principles.setdefault(n, set())
    if c:
        principles[n].add(c)
out["privacy"] = {
    "principles": len(principles),
    "controlLinks": sum(len(v) for v in principles.values()),
}

# ---- Full-content hashes: AOs, ERL, compensating, domains ------------------
ws = sheet_by(r"^assessment objectives")
rows = ws.iter_rows(values_only=True)
hdr = [norm(h) for h in next(rows)]
cctl = next(i for i, h in enumerate(hdr) if re.match(r"^scf #$", h, re.I))
cao = next(i for i, h in enumerate(hdr) if re.match(r"^scf ao #$", h, re.I))
ctext = next(i for i, h in enumerate(hdr) if re.match(r"^scf assessment objective \(ao\)", h, re.I))
crig = next(i for i, h in enumerate(hdr) if re.match(r"^assessment rigor", h, re.I))
corig = next(i for i, h in enumerate(hdr) if re.match(r"^scf assessment objective \(ao\) origin", h, re.I))
ao_rows = []
for r in rows:
    aid = norm(r[cao])
    if not aid:
        continue
    try:
        rig = str(int(float(r[crig])))
        if rig == "0":
            rig = ""
    except (TypeError, ValueError):
        rig = ""
    ao_rows.append("\x1f".join([aid, norm(r[cctl]), norm(r[ctext]), rig, ";".join(split_multi(r[corig]))]))
out["aos"]["contentHash"] = md5("\n".join(sorted(ao_rows)))

ws = sheet_by(r"^evidence request list")
rows = ws.iter_rows(values_only=True)
hdr = [norm(h) for h in next(rows)]
ci_ = hdr.index("ERL #")
ca_ = hdr.index("Area of Focus")
cart = hdr.index("Documentation Artifact")
cd_ = hdr.index("Artifact Description")
cm_ = hdr.index("SCF Control Mappings")
erl_rows = []
for r in rows:
    eid = norm(r[ci_])
    if not eid:
        continue
    erl_rows.append("\x1f".join([eid, norm(r[ca_]), norm(r[cart]), norm(r[cd_]), ";".join(split_multi(r[cm_]))]))
out["erl"]["contentHash"] = md5("\n".join(sorted(erl_rows)))

ws = sheet_by(r"^compensating controls")
rows = ws.iter_rows(values_only=True)
hdr = [norm(h) for h in next(rows)]
cctl = hdr.index("SCF Control #")
crisk = next(i for i, h in enumerate(hdr) if re.match(r"^risk if primary control", h, re.I))
groups = [i for i, h in enumerate(hdr) if re.match(r"^possible compensating control #\d", h, re.I)]
comp_rows = []
for r in rows:
    cid_v = norm(r[cctl])
    if not re.match(r"^[A-Z]{2,4}-\d", cid_v):
        continue
    opts = []
    for g in groups:
        oid = norm(r[g + 2]) if g + 2 < len(r) else ""
        if oid and not re.match(r"^n/?a$", oid, re.I):
            opts.append(f"{oid}:{norm(r[g + 1])}:{norm(r[g + 3]) if g + 3 < len(r) else ''}")
    comp_rows.append("\x1f".join([cid_v, norm(r[crisk]), "|".join(opts)]))
out["compensating"]["contentHash"] = md5("\n".join(sorted(comp_rows)))

ws = sheet_by(r"domains & principles")
rows = list(ws.iter_rows(values_only=True))
hdr = [norm(h) for h in rows[0]]
ci_ = hdr.index("SCF Identifier")
cn_ = hdr.index("SCF Domain")
cp_ = next(i for i, h in enumerate(hdr) if h.endswith("Principles"))
cin_ = hdr.index("Principle Intent")
dom_rows = []
for r in rows[1:]:
    if norm(r[ci_]):
        dom_rows.append("\x1f".join([norm(r[ci_]), norm(r[cn_]), norm(r[cp_]), norm(r[cin_])]))
out["domains"]["contentHash"] = md5("\n".join(sorted(dom_rows)))

# ---- Risk / Threat catalogs -----------------------------------------------------
for key, pat, grp_hdr, id_hdr in [
    ("risks", r"risk catalog", "risk grouping", "risk #"),
    ("threats", r"threat catalog", "threat grouping", "threat #"),
]:
    ws = sheet_by(pat)
    rows = list(ws.iter_rows(values_only=True))
    hrow = next(i for i, r in enumerate(rows) if norm(r[0]).lower() == grp_hdr)
    hdr = [norm(h).lower() for h in rows[hrow]]
    cid_ = hdr.index(id_hdr)
    cname = next(i for i, h in enumerate(hdr) if h.startswith(id_hdr.split(" ")[0] + "*"))
    cdesc = next(i for i, h in enumerate(hdr) if "description" in h)
    entries = []
    cat_rows = []
    for r in rows[hrow + 1:]:
        rid = norm(r[cid_])
        if not rid:
            continue
        entries.append(rid)
        cat_rows.append("\x1f".join([rid, norm(r[cname]), norm(r[cdesc])]))
    out[key] = {"count": len(entries), "ids": sorted(entries), "contentHash": md5("\n".join(sorted(cat_rows)))}

json.dump(out, sys.stdout, indent=1, sort_keys=True)
print(file=sys.stderr)
print(f"controls={controls} aos={aos} erl={erl_count} sources={len(slugs)}", file=sys.stderr)
